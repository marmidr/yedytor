#
# 2025-01-19
#

import logger

# https://linuxhint.com/read-excel-file-python/
# https://openpyxl.readthedocs.io/en/stable/tutorial.html
import openpyxl

from text_grid import TextGrid

# -----------------------------------------------------------------------------

def __check_row_valid(row_cells: list[str]) -> bool:
    # ignore rows with empty cells 'A,B,C' or cell 'A' with a long horizontal line
    row_valid = (len(row_cells) > 3) and (row_cells[0] or row_cells[1] or row_cells[2])
    row_valid = row_valid and not row_cells[0].startswith("___")
    return row_valid

def read_xlsx_sheet(path: str) -> TextGrid:
    """
    Reads entire sheet 0
    """
    assert path is not None
    logger.info(f"Reading file '{path}'")
    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook(path)
    sheet = wookbook.active
    tg = TextGrid()

    # Iterate the loop to read the cell values
    for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
        row_cells = []
        for cell in row:
            if cell is None:
                cell = ""
            elif isinstance(cell, float) or isinstance(cell, int):
                if isinstance(cell, float) and int(cell) == float(cell):
                    # prevent the conversion of '100' to '100.0'
                    cell = int(cell)
                cell = repr(cell)
            # change multiline cells into single-line
            cell = cell.replace("\n", " ‚èé ")
            row_cells.append(cell.strip())

        if __check_row_valid(row_cells):
            tg.rows_raw().append(row_cells)

    tg.nrows = len(tg.rows_raw())
    tg.ncols = sheet.max_column
    tg.align_number_of_columns()
    return tg
